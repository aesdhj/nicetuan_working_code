import pandas as pd
pd.options.display.max_columns = None
pd.options.display.max_rows = None
import warnings
warnings.filterwarnings('ignore')
import os
from datetime import datetime, timedelta, date
from tqdm import tqdm
import sys
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from win32com.client import Dispatch


def just_open(path):
	xlApp = Dispatch("Excel.Application")
	xlApp.Visible = False
	xlBook = xlApp.Workbooks.Open(path)
	xlBook.Save()
	xlBook.Close()


def change_excel(path):
	wb = openpyxl.load_workbook(path)
	sheet = wb.get_active_sheet()

	sheet.insert_rows(1, 2)
	for i in range(11, sheet.max_column+1):
		sheet[get_column_letter(i) + '1'] = '=LEFT(' + get_column_letter(i) + '3, 4)'
		sheet[get_column_letter(i) + '2'] = '=MID(' + get_column_letter(i) + '3,6,LEN(' + get_column_letter(i) + '3))'
	wb.save(path)
	just_open(path)
	wb = openpyxl.load_workbook(path, data_only=True)
	sheet = wb.get_active_sheet()
	for i in range(11, sheet.max_column+1):
		sheet[get_column_letter(i) + '3'].value = sheet[get_column_letter(i) + '2'].value
		if get_column_letter(i) == 'K' or (sheet[get_column_letter(i) + '1'].value == '缺货数量')*(sheet[get_column_letter(i-1) + '1'].value == '多货数量')*(sheet[get_column_letter(i+1) + '1'].value == '缺货数量'):
			sheet[get_column_letter(i) + '2'].value = 1
		else:
			sheet[get_column_letter(i) + '2'].value = 0
	for i in range(11, sheet.max_column+1):
		if sheet[get_column_letter(i) + '2'].value == 1:
			sheet[get_column_letter(i) + '1'].value = sheet[get_column_letter(i) + '1'].value
		else:
			sheet[get_column_letter(i) + '1'].value = ''
	sheet.delete_rows(2, 1)

	sheet.row_dimensions[2].height = 70

	alignment = Alignment(wrap_text=True, vertical='top')
	i = 1
	for row in sheet.rows:
		if i == 2:
			for cell in row:
				cell.alignment = alignment
			break
		else:
			pass
		i += 1

	for i in range(3, sheet.max_column+1):
		if i>=3 and i<=10:
			sheet.column_dimensions[get_column_letter(i)].width = 5
		else:
			sheet.column_dimensions[get_column_letter(i)].width = 1.5

	sheet.freeze_panes = 'A3'

	font = Font(name='微软雅黑', size=10, bold=False)
	for row in sheet.rows:
		for cell in row:
			cell.font = font
			if cell.value == 0:
				cell.value = ''

	wb.save(path)

file_list_14days = [
					(datetime.today() - timedelta(days=i)).strftime('%Y-%m-%d')+'_jiangsu.csv' for i in range(16)] + \
					[(datetime.today() - timedelta(days=i)).strftime('%Y-%m-%d')+'_wenzhou.csv' for i in range(16)] + \
					[(datetime.today() - timedelta(days=i)).strftime('%Y-%m-%d')+'_hefei.csv' for i in range(16)] + \
					[(datetime.today() - timedelta(days=i)).strftime('%Y-%m-%d')+'_hanzhou.csv' for i in range(16)] + \
					[(datetime.today() - timedelta(days=i)).strftime('%Y-%m-%d')+'_xuzhou.csv' for i in range(16)]
path = r'D:\文档\工作\十荟团\temp\华东TC缺货改期_新'
file_list = os.listdir(path)
file_df = []
for file_nm in tqdm(file_list):
	if file_nm in file_list_14days:
		temp_df = pd.read_csv(os.path.join(path, file_nm))
		temp_df['日期'] = file_nm.split('.')[0].split('_')[0]
		if file_nm.split('.')[0].split('_')[1] == 'jiangsu':
			temp_df['主仓'] = '溧阳仓'
			temp_df['主站名称'] = '江苏十荟团'
		elif file_nm.split('.')[0].split('_')[1] == 'wenzhou':
			temp_df['主仓'] = '温州仓'
			temp_df['主站名称'] = '浙南十荟团'
		elif file_nm.split('.')[0].split('_')[1] == 'hefei':
			temp_df['主仓'] = '合肥仓'
			temp_df['主站名称'] = '安徽十荟团'
		elif file_nm.split('.')[0].split('_')[1] == 'xuzhou':
			temp_df['主仓'] = '徐州仓'
			temp_df['主站名称'] = '徐州十荟团'
		else:
			temp_df['主仓'] = '杭州仓'
			temp_df['主站名称'] = '杭州市'
		file_df.append(temp_df)
tc_short_supply_detail = pd.concat(file_df, axis=0, ignore_index=True)

tc_short_supply_detail['日期'] = pd.to_datetime(tc_short_supply_detail['日期'])
tc_short_supply_detail['TC应出库总件数'] = tc_short_supply_detail['TC应出库总件数'].astype(int)
tc_short_supply_detail['多货数量'] = tc_short_supply_detail['多货数量'].astype(int)
tc_short_supply_detail['缺货数量'] = tc_short_supply_detail['缺货数量'].astype(int)
tc_short_supply_detail['缺货数量_TC'] = tc_short_supply_detail['缺货数量_TC'].fillna(0)
tc_short_supply_detail['缺货数量_FC'] = tc_short_supply_detail['缺货数量_FC'].fillna(0)
tc_short_supply_detail['缺货数量_TC'] = tc_short_supply_detail['缺货数量_TC'].astype(int)
tc_short_supply_detail['缺货数量_FC'] = tc_short_supply_detail['缺货数量_FC'].astype(int)
tc_short_supply_detail['破损'] = tc_short_supply_detail['破损'].astype(int)
tc_short_supply_detail['主仓_规格ID_日期'] =tc_short_supply_detail['主仓'].astype(str) + '_' + tc_short_supply_detail['规格ID'].astype(str) + '_' + tc_short_supply_detail['日期'].astype(str)

warehouse_short_supply_detail = pd.read_csv(r'D:\文档\工作\十荟团\temp\缺货改期明细统计.csv', encoding='gbk')
warehouse_short_supply_detail['日期'] = pd.to_datetime(warehouse_short_supply_detail['日期'])
warehouse_short_supply_detail = warehouse_short_supply_detail[(warehouse_short_supply_detail['日期'] <= (date.today() - timedelta(days=0))) & (warehouse_short_supply_detail['日期'] >= (date.today() - timedelta(days=14)))]
warehouse_short_supply_detail = warehouse_short_supply_detail[
	(warehouse_short_supply_detail['来源'] == 'TC') |
	(warehouse_short_supply_detail['来源'] == '安徽LTC') |
	(warehouse_short_supply_detail['来源'] == '无锡LTC') |
	(warehouse_short_supply_detail['来源'] == '武进LTC')]
warehouse_short_supply_detail['规格id'] = warehouse_short_supply_detail['规格id'].astype(int)
warehouse_short_supply_detail['规格id'] = warehouse_short_supply_detail['规格id'].astype(str)
warehouse_short_supply_detail['主仓_规格id_日期'] =warehouse_short_supply_detail['城市仓'].astype(str) + '_' + warehouse_short_supply_detail['规格id'].astype(str) + '_' + warehouse_short_supply_detail['日期'].astype(str)
warehouse_short_supply_list = list(set(warehouse_short_supply_detail['主仓_规格id_日期']))
tc_short_supply_detail['仓库缺货改期商品'] = tc_short_supply_detail.apply(lambda x: 1 if x['主仓_规格ID_日期'] in warehouse_short_supply_list else 0, axis=1)
tc_short_supply_detail['缺货数量_TC'] = tc_short_supply_detail.apply(lambda x: 0 if x['仓库缺货改期商品'] == 1 else x['缺货数量_TC'], axis=1)
tc_short_supply_detail['缺货数量_FC'] = tc_short_supply_detail.apply(lambda x: 0 if x['仓库缺货改期商品'] == 1 else x['缺货数量_FC'], axis=1)
# 2021-07-01
# tc_short_supply_detail['城市圈'] = tc_short_supply_detail.apply(lambda x: x['TC'].split('-')[0], axis=1)
# tc_short_supply_detail = tc_short_supply_detail[tc_short_supply_detail['城市圈'].isin(['安徽', '常州', '南京', '无锡', '浙南', '杭州'])]
tc_short_supply_detail['作业区'] = tc_short_supply_detail.apply(lambda x: '直配' if x['TC'] in ['杭州-嘉兴LTC', '杭州市区LTC', '杭州市区直配', '杭州市区直配', '杭州-嘉兴直配', '新北LTC', '溧阳LTC', '上海嘉定LTC', '虚拟TC'] else 'TC', axis=1)
tc_short_supply_detail.to_csv(r'D:\文档\工作\十荟团\temp\temp_tc_short_supply_detail.csv', index=False, encoding='utf_8_sig')

tc_short_supply_detail['日期'] = pd.to_datetime(tc_short_supply_detail['日期'])
tc_short_supply_check = tc_short_supply_detail[tc_short_supply_detail['日期'] == (date.today() - timedelta(days=0))]
# 2021-07-01
tc_short_supply_check = tc_short_supply_check[tc_short_supply_detail['作业区'] == 'TC']
tc_short_supply_check['规格ID'] = tc_short_supply_check['规格ID'].astype(int)

# liyang
tc_short_supply_check_liyang = tc_short_supply_check[(tc_short_supply_check['主站名称'] == '江苏十荟团')]
# tc_short_supply_check_liyang['TC'] = tc_short_supply_check_liyang.apply(lambda x: x['TC'].split('-')[1], axis=1)
tc_short_supply_check_liyang['仓库缺货改期商品'] = tc_short_supply_check_liyang['仓库缺货改期商品'].astype(str)
total_check_liyang = tc_short_supply_check_liyang.groupby(['规格ID', '商品', '分类', '仓库缺货改期商品'], as_index=False)['TC应出库总件数', '多货数量', '缺货数量', '缺货数量_FC', '缺货数量_TC', '破损',].sum()
tc_short_supply_check_liyang = tc_short_supply_check_liyang.drop(['日期', '主仓_规格ID_日期', 'TC应出库总件数', '破损', '主仓'], axis=1)
tc_short_supply_check_liyang = tc_short_supply_check_liyang.groupby(['规格ID', '商品', '分类', '仓库缺货改期商品', 'TC'], as_index=False)['多货数量', '缺货数量'].sum()
tc_short_supply_check_liyang = tc_short_supply_check_liyang.set_index(['规格ID', '商品', '分类', '仓库缺货改期商品', 'TC']).unstack().reset_index()
tc_short_supply_check_liyang = total_check_liyang.merge(tc_short_supply_check_liyang, on=['规格ID', '商品', '分类', '仓库缺货改期商品'], how='left')
tc_short_supply_check_liyang['仓库缺货改期商品'] = tc_short_supply_check_liyang['仓库缺货改期商品'].astype(int)
tc_short_supply_check_liyang = tc_short_supply_check_liyang[~((tc_short_supply_check_liyang['多货数量'] == 0) & (tc_short_supply_check_liyang['缺货数量'] == 0))]
cols = []
for c in tc_short_supply_check_liyang.columns:
	if ('缺货数量' in c or '多货数量' in c) and (len(c) == 2):
		cols.append(c[0] + '-' + c[1])
	else:
		cols.append(c)
tc_short_supply_check_liyang.columns = cols
path = 'C:/Users/aesdhj/Desktop/江苏_{}_到仓核货差异.xlsx'.format(datetime.now().strftime('%Y%m%d'))
tc_short_supply_check_liyang.to_excel(path, sheet_name='tc_short_supply_check_liyang', index=False)
change_excel(path)

# #wenzhou
# tc_short_supply_check_wenzhou = tc_short_supply_check[(tc_short_supply_check['主站名称'] == '浙南十荟团')]
# # tc_short_supply_check_wenzhou['TC'] = tc_short_supply_check_wenzhou.apply(lambda x: x['TC'].split('-')[1], axis=1)
# tc_short_supply_check_wenzhou['仓库缺货改期商品'] = tc_short_supply_check_wenzhou['仓库缺货改期商品'].astype(str)
# total_check_wenzhou = tc_short_supply_check_wenzhou.groupby(['规格ID', '商品', '分类', '仓库缺货改期商品'], as_index=False)['TC应出库总件数', '多货数量', '缺货数量', '缺货数量_FC', '缺货数量_TC', '破损',].sum()
# tc_short_supply_check_wenzhou = tc_short_supply_check_wenzhou.drop(['日期', '主仓_规格ID_日期', 'TC应出库总件数', '破损', '主仓'], axis=1)
# tc_short_supply_check_wenzhou = tc_short_supply_check_wenzhou.groupby(['规格ID', '商品', '分类', '仓库缺货改期商品', 'TC'], as_index=False)['多货数量', '缺货数量'].sum()
# tc_short_supply_check_wenzhou = tc_short_supply_check_wenzhou.set_index(['规格ID', '商品', '分类', '仓库缺货改期商品', 'TC']).unstack().reset_index()
# tc_short_supply_check_wenzhou = total_check_wenzhou.merge(tc_short_supply_check_wenzhou, on=['规格ID', '商品', '分类', '仓库缺货改期商品'], how='left')
# tc_short_supply_check_wenzhou['仓库缺货改期商品'] = tc_short_supply_check_wenzhou['仓库缺货改期商品'].astype(int)
# tc_short_supply_check_wenzhou = tc_short_supply_check_wenzhou[~((tc_short_supply_check_wenzhou['多货数量'] == 0) & (tc_short_supply_check_wenzhou['缺货数量'] == 0))]
# cols = []
# for c in tc_short_supply_check_wenzhou.columns:
# 	if ('缺货数量' in c or '多货数量' in c) and (len(c) == 2):
# 		cols.append(c[0] + '-' + c[1])
# 	else:
# 		cols.append(c)
# tc_short_supply_check_wenzhou.columns = cols
# path = 'C:/Users/aesdhj/Desktop/浙南_{}_到仓核货差异.xlsx'.format(datetime.now().strftime('%Y%m%d'))
# tc_short_supply_check_wenzhou.to_excel(path, sheet_name='tc_short_supply_check_wenzhou', index=False)
# change_excel(path)

# hefei
# tc_short_supply_check_hanzhou = tc_short_supply_check[(tc_short_supply_check['主站名称'] == '安徽十荟团')]
# # tc_short_supply_check_hanzhou['TC'] = tc_short_supply_check_hanzhou.apply(lambda x: x['TC'].split('-')[1], axis=1)
# tc_short_supply_check_hanzhou['仓库缺货改期商品'] = tc_short_supply_check_hanzhou['仓库缺货改期商品'].astype(str)
# total_check_hanzhou = tc_short_supply_check_hanzhou.groupby(['规格ID', '商品', '分类', '仓库缺货改期商品'], as_index=False)['TC应出库总件数', '多货数量', '缺货数量', '缺货数量_FC', '缺货数量_TC', '破损',].sum()
# tc_short_supply_check_hanzhou = tc_short_supply_check_hanzhou.drop(['日期', '主仓_规格ID_日期', 'TC应出库总件数', '破损', '主仓'], axis=1)
# tc_short_supply_check_hanzhou = tc_short_supply_check_hanzhou.groupby(['规格ID', '商品', '分类', '仓库缺货改期商品', 'TC'], as_index=False)['多货数量', '缺货数量'].sum()
# tc_short_supply_check_hanzhou = tc_short_supply_check_hanzhou.set_index(['规格ID', '商品', '分类', '仓库缺货改期商品', 'TC']).unstack().reset_index()
# tc_short_supply_check_hanzhou = total_check_hanzhou.merge(tc_short_supply_check_hanzhou, on=['规格ID', '商品', '分类', '仓库缺货改期商品'], how='left')
# tc_short_supply_check_hanzhou['仓库缺货改期商品'] = tc_short_supply_check_hanzhou['仓库缺货改期商品'].astype(int)
# tc_short_supply_check_hanzhou = tc_short_supply_check_hanzhou[~((tc_short_supply_check_hanzhou['多货数量'] == 0) & (tc_short_supply_check_hanzhou['缺货数量'] == 0))]
# cols = []
# for c in tc_short_supply_check_hanzhou.columns:
# 	if ('缺货数量' in c or '多货数量' in c) and (len(c) == 2):
# 		cols.append(c[0] + '-' + c[1])
# 	else:
# 		cols.append(c)
# tc_short_supply_check_hanzhou.columns = cols
# path = 'C:/Users/aesdhj/Desktop/安徽_{}_到仓核货差异.xlsx'.format(datetime.now().strftime('%Y%m%d'))
# tc_short_supply_check_hanzhou.to_excel(path, sheet_name='tc_short_supply_check_hefei', index=False)
# change_excel(path)

# xuzhou
# tc_short_supply_check_hanzhou = tc_short_supply_check[(tc_short_supply_check['主站名称'] == '徐州十荟团')]
# # tc_short_supply_check_hanzhou['TC'] = tc_short_supply_check_hanzhou.apply(lambda x: x['TC'].split('-')[1], axis=1)
# tc_short_supply_check_hanzhou['仓库缺货改期商品'] = tc_short_supply_check_hanzhou['仓库缺货改期商品'].astype(str)
# total_check_hanzhou = tc_short_supply_check_hanzhou.groupby(['规格ID', '商品', '分类', '仓库缺货改期商品'], as_index=False)['TC应出库总件数', '多货数量', '缺货数量', '缺货数量_FC', '缺货数量_TC', '破损',].sum()
# tc_short_supply_check_hanzhou = tc_short_supply_check_hanzhou.drop(['日期', '主仓_规格ID_日期', 'TC应出库总件数', '破损', '主仓'], axis=1)
# tc_short_supply_check_hanzhou = tc_short_supply_check_hanzhou.groupby(['规格ID', '商品', '分类', '仓库缺货改期商品', 'TC'], as_index=False)['多货数量', '缺货数量'].sum()
# tc_short_supply_check_hanzhou = tc_short_supply_check_hanzhou.set_index(['规格ID', '商品', '分类', '仓库缺货改期商品', 'TC']).unstack().reset_index()
# tc_short_supply_check_hanzhou = total_check_hanzhou.merge(tc_short_supply_check_hanzhou, on=['规格ID', '商品', '分类', '仓库缺货改期商品'], how='left')
# tc_short_supply_check_hanzhou['仓库缺货改期商品'] = tc_short_supply_check_hanzhou['仓库缺货改期商品'].astype(int)
# tc_short_supply_check_hanzhou = tc_short_supply_check_hanzhou[~((tc_short_supply_check_hanzhou['多货数量'] == 0) & (tc_short_supply_check_hanzhou['缺货数量'] == 0))]
# cols = []
# for c in tc_short_supply_check_hanzhou.columns:
# 	if ('缺货数量' in c or '多货数量' in c) and (len(c) == 2):
# 		cols.append(c[0] + '-' + c[1])
# 	else:
# 		cols.append(c)
# tc_short_supply_check_hanzhou.columns = cols
# path = 'C:/Users/aesdhj/Desktop/徐州_{}_到仓核货差异.xlsx'.format(datetime.now().strftime('%Y%m%d'))
# tc_short_supply_check_hanzhou.to_excel(path, sheet_name='tc_short_supply_check_xuzhou', index=False)
# change_excel(path)

# hanzhou
# tc_short_supply_check_hanzhou = tc_short_supply_check[(tc_short_supply_check['主站名称'] == '杭州市')]
# # tc_short_supply_check_hanzhou['TC'] = tc_short_supply_check_hanzhou.apply(lambda x: x['TC'].split('-')[1], axis=1)
# tc_short_supply_check_hanzhou['仓库缺货改期商品'] = tc_short_supply_check_hanzhou['仓库缺货改期商品'].astype(str)
# total_check_hanzhou = tc_short_supply_check_hanzhou.groupby(['规格ID', '商品', '分类', '仓库缺货改期商品'], as_index=False)['TC应出库总件数', '多货数量', '缺货数量', '缺货数量_FC', '缺货数量_TC', '破损',].sum()
# tc_short_supply_check_hanzhou = tc_short_supply_check_hanzhou.drop(['日期', '主仓_规格ID_日期', 'TC应出库总件数', '破损', '主仓'], axis=1)
# tc_short_supply_check_hanzhou = tc_short_supply_check_hanzhou.groupby(['规格ID', '商品', '分类', '仓库缺货改期商品', 'TC'], as_index=False)['多货数量', '缺货数量'].sum()
# tc_short_supply_check_hanzhou = tc_short_supply_check_hanzhou.set_index(['规格ID', '商品', '分类', '仓库缺货改期商品', 'TC']).unstack().reset_index()
# tc_short_supply_check_hanzhou = total_check_hanzhou.merge(tc_short_supply_check_hanzhou, on=['规格ID', '商品', '分类', '仓库缺货改期商品'], how='left')
# tc_short_supply_check_hanzhou['仓库缺货改期商品'] = tc_short_supply_check_hanzhou['仓库缺货改期商品'].astype(int)
# tc_short_supply_check_hanzhou = tc_short_supply_check_hanzhou[~((tc_short_supply_check_hanzhou['多货数量'] == 0) & (tc_short_supply_check_hanzhou['缺货数量'] == 0))]
# cols = []
# for c in tc_short_supply_check_hanzhou.columns:
# 	if ('缺货数量' in c or '多货数量' in c) and (len(c) == 2):
# 		cols.append(c[0] + '-' + c[1])
# 	else:
# 		cols.append(c)
# tc_short_supply_check_hanzhou.columns = cols
# path = 'C:/Users/aesdhj/Desktop/杭州_{}_到仓核货差异.xlsx'.format(datetime.now().strftime('%Y%m%d'))
# tc_short_supply_check_hanzhou.to_excel(path, sheet_name='tc_short_supply_check_hanzhou', index=False)
# change_excel(path)

sys.exit(0)
